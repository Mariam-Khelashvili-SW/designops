"""Design intake generator — paste handover email → LLM → preview → Notion publish."""

from __future__ import annotations

import csv
import io
import logging
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from sqlalchemy.orm import Session

from designops.adapters.llm import LLMClient, LLMResult, parse_digest_json
from designops.adapters.notion import NotionClient
from designops.core.models import IntakeDraft, Pipeline

log = logging.getLogger(__name__)

PIPELINE_KEY = "design-intake"
SKILL_PATH = Path(__file__).resolve().parents[1] / "skills" / "design-intake.md"

SAMPLE_INTAKE = {
    "pasted_input": """From: iryna.rubanava@scandiweb.com
To: design-ops@scandiweb.com
Cc: olga.kimalana@scandiweb.com, ana.taylor@scandiweb.com
Subject: C-pipe handover — Sports Group Denmark Club Portal (wireframes)

Hi team,

Handing over a new design project from C-pipe.

Client: Sports Group Denmark (SGD)
Project: Club Portal rebuild — wireframes phase
Scope: Core (club-specific screens only; reuse the SGD B2B store design system)
Site type: B2B

Sports Group Denmark makes sportswear. Badminton clubs order team uniforms through a private Club Portal. A club member logs in, sees only their club's products at club prices, picks a jersey, adds their name and the club logo, and orders before the club's shared deadline. The club collects orders into one batch; clubs can pay for members using club credit.

The current portal is old and runs on a lot of manual work. It is on Shopify today. SGD is growing from 30 clubs to 40+ this year, so we are rebuilding the portal on the same Hyva / Magento stack as the SGD web store we already designed — reusing that design system, and designing new only what is unique to clubs.

Website / current Club Portal URL: not in this handover — please request from the client.

Commercial
- Signed for this phase: wireframes, 30–50h
- Two revision rounds included; anything past that is billed extra
- Full estimate: see the sheet (hours below). Proposal slides are linked separately.
- Contract status: signed for Discovery / wireframes. Build-phase Designs not signed yet.

This phase (~4 weeks): wireframes for the club-specific screens so open questions get answered and the client can approve the build. Design leads. Main deliverable = wireframes.

What we design (new)
- Club landing page — each club gets its own page: banner, club info, their products
- Product page — member picks name + club logo, previews it; bundles and a possible sponsor choice
- Cart & checkout — order deadline must be visible; club credit as a payment method
- Sign-up & log-in — new members join a club; someone approves them (approval step)
- My account — standard account + a block showing the club's deadline
- Global elements, homepage, product list — light adjustments of what we already have from the SGD store

Reuse, don't redesign
- Look & feel, colors, type, components — from the existing SGD store design system (Figma link from the SGD B2B project — being added, not in this mail)
- Search, CMS pages, transactional emails — reused with styling touch-ups only

Out of scope
- No public-facing shop, no fan merchandise
- No automatic logo-on-photo rendering (stays manual on the client side)
- No extra screen sizes beyond desktop 1440px and mobile 375px

Open questions (need answers in Discovery)
1. How do club budgets / sponsor money work — who pays for what, and what does the member see at checkout? (Mette owes us this — she is on vacation after kickoff week)
2. When someone signs up, are they approved automatically (e.g. Danish address) or manually by the club admin?
3. Where does the order deadline live — one date for everyone or per club, and who sets it?
4. How different are club prices from the main store's logic? (tech, during Discovery)
5. Invoice currency for the SOW is EUR 18,400 vs the estimate sheet which still shows DKK — do not resolve, just flag. No design impact.

People
- Iryna Rubanava — DM (first stop for scope, deadlines, client communication)
- Olga Kimalana — design lead (direction, wireframe validation, who is assigned). UX/UI designer not assigned yet — Olga to assign
- Ana Taylor — KAM (client relationship, brand asset requests)
- Mette — client, SGD (how clubs work day-to-day; budget/sponsor rules). On vacation after kickoff week

Kickoff: planned this Friday — to confirm.

Constraints
- The portal shares its website foundation with the SGD store that's being built right now (shared release calendar — handoff dates are real dates)
- Member accounts are not migrated: club data moves over via Twoday (middleware vendor), but individual members are re-created by club admins after launch

Missing from handover
- Brand assets / club logo pack — not attached, request via Ana
- WoW (Ways of Working) doc — not provided, please request
- Current Club Portal URL — requested

Figma / Design Drive / Jira: pending, added at design kickoff.

Thanks,
Iryna
""",
    "estimate_link": "https://docs.google.com/spreadsheets/d/sample-sgd-club-portal-estimate/edit",
    "proposal_link": "https://docs.google.com/presentation/d/sample-sgd-club-portal-proposal/edit",
    "estimate_rows": (
        "Phase\tTask\tHours\n"
        "Wireframes\tGlobal elements (header, footer)\t4\n"
        "Wireframes\tHomepage (light adjust)\t3\n"
        "Wireframes\tClub landing page\t6\n"
        "Wireframes\tProduct list (light adjust)\t3\n"
        "Wireframes\tProduct page + name/logo picker\t12\n"
        "Wireframes\tCart\t4\n"
        "Wireframes\tCheckout (deadline + club credit)\t8\n"
        "Wireframes\tSign-up / log-in / password (with approval step)\t6\n"
        "Wireframes\tMy account + deadline block\t4\n"
        "Wireframes\tRevisions (2 rounds)\t8\n"
        "Designs (later phase — do not seed)\tHi-fi UI\t80"
    ),
    "corrections": (
        "KAM is Ana Taylor. Kickoff is planned this Friday — still to confirm. "
        "Designer is unassigned — Olga to assign. Do not invent a Club Portal URL."
    ),
}


@dataclass(slots=True)
class GenerateResult:
    draft_id: uuid.UUID
    title: str | None
    error: str | None


@dataclass(slots=True)
class PublishResult:
    draft_id: uuid.UUID
    notion_page_url: str
    notion_page_id: str


def get_intake_pipeline(session: Session) -> Pipeline:
    pipe = session.query(Pipeline).filter_by(key=PIPELINE_KEY).one_or_none()
    if pipe is None:
        raise RuntimeError("design-intake pipeline not bootstrapped — restart app")
    return pipe


def load_system_prompt() -> str:
    raw = SKILL_PATH.read_text(encoding="utf-8")
    # Strip markdown heading if present; use body as system prompt
    lines = raw.splitlines()
    if lines and lines[0].startswith("#"):
        return "\n".join(lines[1:]).strip()
    return raw.strip()


def parse_spreadsheet(filename: str, content: bytes) -> str:
    """Extract plain text from uploaded csv/xlsx for LLM context."""
    lower = (filename or "").lower()
    if lower.endswith(".csv"):
        text = content.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = ["\t".join(cell for cell in row) for row in reader]
        return "\n".join(rows)
    if lower.endswith((".xlsx", ".xls")):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return ""
        lines: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                lines.append("\t".join(cells))
        wb.close()
        return "\n".join(lines)
    raise ValueError(f"Unsupported file type: {filename} (use .csv or .xlsx)")


def build_user_message(
    *,
    pasted_input: str,
    estimate_link: str = "",
    proposal_link: str = "",
    estimate_rows: str = "",
    uploaded_files: list[dict[str, str]] | None = None,
    corrections: str = "",
) -> str:
    parts = [f"EMAIL_CONTENT:\n{pasted_input.strip()}"]
    if estimate_link.strip():
        parts.append(f"ESTIMATE_LINK:\n{estimate_link.strip()}")
    if proposal_link.strip():
        parts.append(f"PROPOSAL_LINK:\n{proposal_link.strip()}")
    if estimate_rows.strip():
        parts.append(f"ESTIMATE_ROWS:\n{estimate_rows.strip()}")
    for uf in uploaded_files or []:
        fname = uf.get("filename") or "upload"
        text = uf.get("extracted_text") or ""
        if text.strip():
            parts.append(f"UPLOADED_FILE ({fname}):\n{text.strip()}")
    if corrections.strip():
        parts.append(f"CORRECTIONS:\n{corrections.strip()}")
    return "\n\n".join(parts)


def validate_sections_json(data: dict) -> str | None:
    """Return error message if invalid, else None."""
    if not isinstance(data, dict):
        return "LLM output was not a JSON object"
    if data.get("error"):
        return str(data["error"])
    if not data.get("title"):
        return "LLM output missing title"
    return None


def generate_intake_draft(
    session: Session,
    *,
    pasted_input: str,
    estimate_link: str = "",
    proposal_link: str = "",
    estimate_rows: str = "",
    corrections: str = "",
    uploaded_files: list[dict[str, str]] | None = None,
    runner: str | None = None,
    draft_id: uuid.UUID | None = None,
) -> GenerateResult:
    """Generate or regenerate an intake draft via one LLM call."""
    pasted_input = (pasted_input or "").strip()
    if not pasted_input:
        raise ValueError("Handover email content is required")

    user_content = build_user_message(
        pasted_input=pasted_input,
        estimate_link=estimate_link,
        proposal_link=proposal_link,
        estimate_rows=estimate_rows,
        uploaded_files=uploaded_files,
        corrections=corrections,
    )

    llm = LLMClient()
    system = load_system_prompt()
    result: LLMResult | None = None
    data: dict = {}
    parse_err: str | None = None

    for attempt in range(2):
        try:
            result = llm.synthesize(system=system, user_content=user_content, max_tokens=16000)
            data = parse_digest_json(result.text)
            parse_err = validate_sections_json(data)
            if parse_err and attempt == 0 and "JSON" in parse_err:
                continue
            break
        except Exception as e:  # noqa: BLE001
            if attempt == 1:
                raise
            log.warning("intake LLM attempt %s failed: %s", attempt + 1, e)

    if result is None:
        raise RuntimeError("LLM call failed")

    llm_error = data.get("error")
    title = data.get("title")
    intake_report = data.get("intake_report") or ""
    flags = data.get("flags") or []

    if draft_id:
        draft = session.get(IntakeDraft, draft_id)
        if draft is None:
            raise ValueError(f"Draft {draft_id} not found")
    else:
        draft = IntakeDraft(pasted_input=pasted_input)

    draft.pasted_input = pasted_input
    draft.estimate_link = estimate_link or None
    draft.proposal_link = proposal_link or None
    draft.estimate_rows = estimate_rows or None
    draft.corrections = corrections or None
    draft.uploaded_files = uploaded_files or []
    draft.title = title
    draft.sections_json = data
    draft.intake_report = intake_report if isinstance(intake_report, str) else str(intake_report)
    draft.flags = flags if isinstance(flags, list) else []
    draft.runner = runner
    draft.input_tokens = result.input_tokens
    draft.output_tokens = result.output_tokens
    draft.cost_usd = result.cost_usd

    if llm_error or parse_err:
        draft.status = "error"
        draft.error_message = llm_error or parse_err
        session.add(draft)
        session.commit()
        return GenerateResult(draft_id=draft.id, title=title, error=draft.error_message)

    draft.status = "draft"
    draft.error_message = None
    session.add(draft)
    session.commit()
    return GenerateResult(draft_id=draft.id, title=title, error=None)


def publish_intake_draft(session: Session, draft_id: uuid.UUID) -> PublishResult:
    """Publish a reviewed draft to Notion."""
    draft = session.get(IntakeDraft, draft_id)
    if draft is None:
        raise ValueError(f"Draft {draft_id} not found")
    if draft.status == "published" and draft.notion_page_url:
        raise ValueError("Draft already published")
    if draft.status == "error" or not draft.sections_json:
        raise ValueError("Cannot publish — draft has errors or no content")

    sections = draft.sections_json
    if sections.get("error"):
        raise ValueError(f"Cannot publish — {sections['error']}")

    title = draft.title or sections.get("title") or "Untitled intake"
    notion = NotionClient()
    page_id, page_url = notion.create_intake_page(title=title, sections_json=sections)

    draft.notion_page_id = page_id
    draft.notion_page_url = page_url
    draft.status = "published"
    draft.published_at = datetime.now(timezone.utc)
    session.add(draft)
    session.commit()

    return PublishResult(draft_id=draft.id, notion_page_url=page_url, notion_page_id=page_id)


def render_preview_html(sections: dict) -> str:
    """Render sections_json as HTML for admin preview."""
    if not sections:
        return "<p>No content.</p>"
    if sections.get("error"):
        return f'<div class="banner warn">{sections["error"]}</div>'

    parts: list[str] = []

    props = sections.get("properties_callout") or {}
    if props:
        prop_lines = " · ".join(
            f"{k.replace('_', ' ').title()}: {v}"
            for k, v in props.items()
            if v
        )
        parts.append(f'<div class="intake-callout intake-callout-props">📋 {prop_lines}</div>')

    s1 = sections.get("section_1") or {}
    if s1:
        parts.append("<h2>1. What this project is</h2>")
        if s1.get("website"):
            parts.append(f"<p><strong>Website:</strong> {s1['website']}</p>")
        for p in s1.get("paragraphs") or []:
            parts.append(f"<p>{p}</p>")
        if s1.get("job_callout"):
            parts.append(f'<div class="intake-callout intake-callout-blue">🎯 {s1["job_callout"]}</div>')

    s2 = sections.get("section_2") or {}
    if s2:
        parts.append("<h2>2. Who uses it</h2>")
        parts.append('<div class="intake-personas">')
        for p in s2.get("personas") or []:
            parts.append(
                f'<div class="intake-persona"><strong>{p.get("title","")}</strong>'
                f'<p>{p.get("description","")}</p></div>'
            )
        parts.append("</div>")
        if s2.get("not_for"):
            parts.append(f"<p><em>{s2['not_for']}</em></p>")

    s3 = sections.get("section_3") or {}
    if s3:
        h = "3. What we design"
        if s3.get("phase_hours"):
            h += f" ({s3['phase_hours']})"
        parts.append(f"<h2>{h}</h2>")
        screens = s3.get("screens") or []
        if screens:
            parts.append("<p><strong>✏️ New screens</strong></p>")
            parts.append(_html_table(["Screen", "What's special"], screens, "screen", "special"))
        if s3.get("reuse"):
            parts.append("<p><strong>♻️ Reuse</strong></p><ul>")
            for r in s3["reuse"]:
                parts.append(f"<li>{r}</li>")
            parts.append("</ul>")
        if s3.get("out_of_scope"):
            parts.append("<p><strong>🚫 Out of scope</strong></p><ul>")
            for r in s3["out_of_scope"]:
                parts.append(f"<li>{r}</li>")
            parts.append("</ul>")
        if s3.get("revision_callout"):
            parts.append(f'<div class="intake-callout intake-callout-yellow">✂️ {s3["revision_callout"]}</div>')

    s4 = sections.get("section_4") or {}
    if s4:
        parts.append("<h2>4. Open questions — and which screens they block</h2>")
        rows = s4.get("rows") or []
        if rows:
            parts.append(_html_table(["Open question", "Blocks", "Who"], rows, "question", "blocks", "who"))
        if s4.get("closing"):
            parts.append(f"<p>{s4['closing']}</p>")

    s5 = sections.get("section_5") or {}
    if s5:
        parts.append("<h2>5. People</h2>")
        rows = s5.get("rows") or []
        if rows:
            parts.append(_html_table(["Who", "Ask them about"], rows, "who", "ask_about"))

    s6 = sections.get("section_6") or {}
    if s6:
        parts.append("<h2>6. Heads-up</h2><ul>")
        for b in s6.get("bullets") or []:
            parts.append(f"<li>{b}</li>")
        parts.append("</ul>")

    s7 = sections.get("section_7") or {}
    if s7:
        parts.append("<h2>7. Files &amp; links</h2>")
        rows = s7.get("rows") or []
        if rows:
            parts.append(_html_table(["What", "Status", "Where"], rows, "what", "status", "where"))

    s8 = sections.get("section_8") or {}
    if s8:
        parts.append("<h2>8. Working space</h2>")
        parts.append("<p><strong>1️⃣ Kick off meeting</strong></p>")
        parts.append("<p><strong>2️⃣ UX discovery</strong></p>")
        parts.append("<p><strong>3️⃣ UX/UI designs</strong></p>")
        plinks = s8.get("project_links") or []
        if plinks:
            parts.append("<p><strong>Project links</strong></p>")
            parts.append(_html_table(["Link", "Where"], plinks, "label", "url"))
        templates = s8.get("templates") or []
        if templates:
            parts.append("<p><strong>Templates to produce — Wireframes</strong></p>")
            parts.append(_html_table(["Template", "Status"], templates, "name", "status"))
        if s8.get("sync_callout"):
            parts.append(f'<div class="intake-callout intake-callout-gray">🔄 {s8["sync_callout"]}</div>')

    s9 = sections.get("section_9") or {}
    if s9:
        parts.append("<h2>9. Project AI assistant</h2>")
        parts.append(f"<p>🔗 {s9.get('assistant_link', 'paste link here')}</p>")

    s10 = sections.get("section_10") or {}
    if s10:
        parts.append("<h2>10. Case study</h2>")
        parts.append(f"<p>{s10.get('placeholder', '')}</p>")

    return "\n".join(parts)


def _html_table(headers: list[str], rows: list[dict], *keys: str) -> str:
    out = ["<table><thead><tr>"]
    for h in headers:
        out.append(f"<th>{h}</th>")
    out.append("</tr></thead><tbody>")
    for row in rows:
        out.append("<tr>")
        for k in keys:
            out.append(f"<td>{row.get(k, '')}</td>")
        out.append("</tr>")
    out.append("</tbody></table>")
    return "".join(out)
